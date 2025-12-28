import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import seaborn as sns
from sklearn import datasets

def create_correlation_heatmap(df, diagrams_folder_path, title= None, filename = None):
    corr = df.select_dtypes('number').corr()
    if title is not None:
        plt.title(title)
    sns.heatmap(data=corr,
                vmin=-1,
                vmax=1,
                center=0)
    if filename is None:
        filename = 'heatmap.png'
    plt.savefig(f'{diagrams_folder_path}/{filename}', bbox_inches='tight')
    plt.clf()

    print("Created correlation heatmap")


def create_pairplot(df, diagrams_folder_path, hue_column=None):
    if hue_column is None:
        sns.pairplot(df.select_dtypes('number'))
    else:
        sns.pairplot(df.select_dtypes('number'))
    plt.savefig(f'{diagrams_folder_path}/pairplot.png', bbox_inches='tight')
    plt.clf()
    plt.figure(figsize=(8, 6))

def save_correlation_heatmap_to_excel(wb, diagram_folder_path, filename=None):
    ws = wb['data_audit']
    if filename is None:
        filename = 'heatmap.png'
    img = openpyxl.drawing.image.Image(f'{diagram_folder_path}/{filename}')
    row = ws.max_row + 1
    col = 'A'
    cell_ref = f"{col}{row}"
    
    img.anchor = cell_ref
    ws.add_image(img)
    print("Saved correlation heatmap in Excel")

def save_pairplot_to_excel(wb, diagram_folder_path):
    ws = wb['data_audit']
    img = openpyxl.drawing.image.Image(f'{diagram_folder_path}/pairplot.png')
    row = ws.max_row + 1
    col = 'A'
    cell_ref = f"{col}{row}"
    
    img.anchor = cell_ref
    ws.add_image(img)

    print("Saved pairplot in Excel")

def create_bar_or_histogram_chart_for_column_and_save_data_to_excel(df, col, bar_chart_cols, excel_writer, diagram_folder_path, create_chart = True):
    value_counts = df[col].value_counts().sort_values(ascending=False)
    value_counts.T.to_excel(excel_writer=excel_writer, sheet_name=col)
    if create_chart:
        if col in bar_chart_cols:
            sns.barplot(value_counts)
            plt.title(f'{col} bar chart')
        else:
            sns.histplot(df[col])
            plt.title(f'{col} histogram')
            plt.ylabel('count')

        plt.xlabel(col)
        plt.xticks(rotation=90, ha='left')

        plt.savefig(f'{diagram_folder_path}/{col}.png')
        plt.cla()

    print("Saved column data to Excel")
    print(f"Created histograms and bar charts for column {col}")

def save_column_chart_to_excel(wb, col, diagram_folder_path):
    ws = wb[col]
    img = openpyxl.drawing.image.Image(f'{diagram_folder_path}/{col}.png')
    img.anchor = 'D1'
    ws.add_image(img)

def describe_df(df):
    desc = df.describe(include='all').T
    desc['num_unique']        = df.nunique()
    desc['pct_unique']        = df.nunique() / len(df) * 100
    desc['num_missing']       = df.isna().sum()
    desc['pct_missing']       = df.isna().sum() / len(df) * 100

    desc = desc.round(2)
    return desc

def create_data_audit(df, filename = './data_audit.xlsx', diagrams_folder_path = './diagrams', bar_chart_cols = [], no_chart_cols = [], hue_column=None, create_pairplot_ = True, create_sheets = True):
    excel_writer = pd.ExcelWriter(filename)

    describe_df(df).to_excel(excel_writer=excel_writer, sheet_name='data_audit')

    create_correlation_heatmap(df, diagrams_folder_path)
    if create_pairplot_:
        create_pairplot(df, diagrams_folder_path, hue_column=hue_column)

    if create_sheets:
        for col in df.columns:
            if col == 'Unnamed: 0':
                continue

            if col in no_chart_cols:
                create_bar_or_histogram_chart_for_column_and_save_data_to_excel(df, col, bar_chart_cols, excel_writer, diagrams_folder_path, create_chart = False)
            else:
                create_bar_or_histogram_chart_for_column_and_save_data_to_excel(df, col, bar_chart_cols, excel_writer, diagrams_folder_path)


    excel_writer.close()

    wb = openpyxl.load_workbook(filename)
    save_correlation_heatmap_to_excel(wb, diagrams_folder_path)
    if create_pairplot_:
        save_pairplot_to_excel(wb, diagrams_folder_path)

    if create_sheets:
        for col in df.columns:
            if col == 'Unnamed: 0' or col in no_chart_cols:
                continue
            save_column_chart_to_excel(wb, col, diagrams_folder_path)

    print("Saved column charts in Excel")
    wb.save(filename)
