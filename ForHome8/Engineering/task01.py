import pandas as pd
import numpy as np

import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parents[2]))

import Homework1.clusterization as clusterization
import openpyxl
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE
from sklearn.pipeline import Pipeline

from sklearn.cluster import DBSCAN
import hdbscan
from sklearn.preprocessing import StandardScaler
import numpy as np
from sklearn.pipeline import Pipeline
from ml_lib.hierarchy import AgglomerativeClustering
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from sklearn.metrics import silhouette_score
from scipy.cluster.hierarchy import dendrogram, linkage
from sklearn.model_selection import RandomizedSearchCV
from sklearn.model_selection import PredefinedSplit
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE
from sklearn.decomposition import TruncatedSVD
from scipy.sparse import csr_matrix


def clusterize(X, name, y=None, n_features = 2, scaling=True, pca = True, tsne = True, truncatedsvd = False, args = None, export_to_file = True, ws = None, diagrams_dir='./diagrams', real_labels = None, clusterization_distance = 1):
    base_model_silhouette = -1

    print(args)
    param_grid = None
    reg = None
    
    if scaling == True:
        steps = [('scaler', StandardScaler())]
    else:
        steps = []

    if pca == True:
        steps.append(('pca', PCA(n_components=n_features)))

    # if tsne == True:
    #     steps.append(('tsne', TSNE(n_components=n_features)))

    if truncatedsvd == True:
        steps.append(('truncatedsvd', TruncatedSVD(n_components=n_features)))

    if 'hdbscan' in name:
        if args == None:
            param_grid = {
                # 'hdbscan__min_cluster_size': np.arange(1, 50, 1)
                # 'hdbscan__min_cluster_size': [2, 1000, 5000, 10000, 20000]
                 'hdbscan__min_cluster_size': [5000]
            }
        else:
            param_grid = args['params']
        reg = hdbscan.HDBSCAN()
        steps.append(('hdbscan', reg))
    elif 'dbscan' in name:
        if args == None:
            param_grid = {
                'dbscan__eps': np.arange(.1, clusterization_distance, .05*clusterization_distance),
                'dbscan__min_samples': np.arange(1, 50, 1)
                }
        else:
            param_grid = args['params']
        reg = DBSCAN()
        steps.append(('dbscan', reg))
    elif 'agglomerative' in name:
        if args == None:
            param_grid = {
                'agglomerative__linkage': ['single', 'average', 'complete'],
                'agglomerative__distance_threshold':np.arange(.1, clusterization_distance, .05*clusterization_distance)
                }
        else:
            param_grid = args['params']
        reg = AgglomerativeClustering(
            distance_threshold=1
        )

        steps.append(('agglomerative', reg))


    pipeline = Pipeline(steps)

    
    cv = [(np.arange(X.shape[0]), np.arange(X.shape[0]))]

    cv = RandomizedSearchCV(
        pipeline,
        param_distributions=param_grid,
        scoring=clusterization.silhouette_scorer,
        cv=cv,
        n_iter=30
    )

    cv.fit(X=X)

    if export_to_file:
        pipeline = cv.best_estimator_

        clusterer = pipeline.named_steps[list(pipeline.named_steps.keys())[-1]]
        labels = pipeline.fit_predict(X)

        # Calculate metrics
        inertia = getattr(clusterer, "inertia_", None)  # only for KMeans / MiniBatchKMeans
        silhouette = silhouette_score(X, labels) if len(set(labels)) > 1 else None

        scatter_path_test = None
        silhouette_test = None

        if y is not None:
            labels_y = pipeline.fit_predict(y)
            silhouette_test = silhouette_score(y, labels_y) if len(set(labels_y)) > 1 else None
        
            # ----------- SCATTERPLOT OF CLUSTERS -----------
            plt.figure(figsize=(7, 5))
            plt.scatter(y.iloc[:, 0], y.iloc[:, 1], c=labels_y)
            plt.title(f"{name} — Cluster Test Scatterplot")
            plt.xlabel(y.columns[0])
            plt.ylabel(y.columns[1])
            plt.tight_layout()

            scatter_path_test = f"{diagrams_dir}/{name}_scatter_test.png"
            plt.savefig(scatter_path_test)
            plt.cla()
        
        if not isinstance(X, csr_matrix):
            if X.shape[1] > 2:
                pca = PCA(n_components=2)
                X_2d = pca.fit_transform(X)
            else:
                X_2d = np.array(X)
            # ----------- SCATTERPLOT OF CLUSTERS -----------
            plt.figure(figsize=(7, 5))
            plt.scatter(X_2d[:, 0], X_2d[:, 1], c=labels)
            plt.title(f"{name} — Cluster Scatterplot")

            plt.tight_layout()

            scatter_path = f"{diagrams_dir}/{name}_scatter.png"
            plt.savefig(scatter_path)
            plt.cla()

        

        # ----------- DENDROGRAM (ONLY FOR HIERARCHICAL) -----------
        dendro_path = None
        if clusterer.__class__.__name__.startswith("Agglomerative"):
            distance_threshold = clusterer.distance_threshold
           
            plt.figure(figsize=(10, 5))
            Z = linkage(X, method=clusterer.linkage)
            dendrogram(Z,labels=X.index, color_threshold=distance_threshold)
            plt.title(f"{name} — Dendrogram ({clusterer.linkage} linkage)")
            plt.tight_layout()

            dendro_path = f"{diagrams_dir}/{name}_dendrogram.png"
            plt.savefig(dendro_path)
            plt.cla()

        pca_vairance_path = None
        if pca == True:
            pca = pipeline.named_steps["pca"]

            variances = pca.explained_variance_

            plt.figure(figsize=(8, 5))
            plt.bar(np.arange(1, len(variances)+1), variances)
            plt.xlabel("Principal Component")
            plt.ylabel("Explained Variance")
            plt.title("PCA: Explained Variance per Component")
            plt.tight_layout()
            
            pca_vairance_path = f"{diagrams_dir}/{name}_pca_variance.png"
            plt.savefig(pca_vairance_path)
            plt.cla()
        
        crosstab_path = None
        if real_labels is not None:
            crosstab_path = f"{diagrams_dir}/{name}_crosstab_output.png"
            cross_tab = pd.crosstab(real_labels, labels, rownames=["Actual"], colnames=["Predicted"])
            clusterization.save_crosstab_image(cross_tab, crosstab_path)

        # ----------- EXPORT TO EXCEL -----------
        col = 'J'
        for img_path in [scatter_path, dendro_path, scatter_path_test, pca_vairance_path, crosstab_path]:
            if not img_path:
                col = chr(ord(col) + 1)
                continue

            img = openpyxl.drawing.image.Image(img_path)
            row = ws.max_row + 1
            cell_ref = f"{col}{row}"

            img_width, img_height = img.width, img.height
            ws.column_dimensions[col].width = img_width / 7
            ws.row_dimensions[row].height = img_height

            img.anchor = cell_ref
            ws.add_image(img)

            col = chr(ord(col) + 1)

        # ----------- APPEND SUMMARY ROW -----------
        ws.append([
            f"{name} Clustering",
            str((scaling, pca, tsne, truncatedsvd)),
            X.shape[1],
            str(cv.best_params_) ,
            inertia if inertia is not None else "",
            silhouette if silhouette is not None else "",
            silhouette/base_model_silhouette * 100 - 100 if silhouette is not None else 0,
            silhouette_test if silhouette_test is not None else "",
            silhouette_test/base_model_silhouette * 100 - 100 if silhouette_test is not None else 0,


        ])

    return pipeline

def main():
    X = pd.read_csv("./ForHome8/DataScience/seeds_dataset.txt", 
                 sep=r"\s+",        # split on whitespace / tabs
                 header=None,       # no header in your snippet
                 engine="python",   # allows regex separators
                 names = ['area', 'perimeter', 'compactness', 'kernel_length', 'kernel_width', 'assymetry_coef', 'kernel_groove_length', 'varieties']
                 )  
    
    X['varieties'] -= 1

    varieties_mapping = {0:'Kama wheat', 1:'Rosa wheat', 2:'Canadian wheat'}
    real_labels = X['varieties'].map(varieties_mapping)

    filename = './ForHome8/Engineering/model_report_task01_engineering.xlsx'
    diagrams_folder_path = './ForHome8/DataScience/diagrams'

    wb = openpyxl.Workbook()
    wb.create_sheet('ModelReport')
    ws = wb['ModelReport']

    ws.append(['Model', 'Scaling/PCA/t-SNE/TruncatedSVD', 'Number of variables','Hyperparams', 'Inertia', 'Silhouette', 'Silhouette Increase from base model %', 'Silhouette for test data', 'Silhouette for test data increase %','Scatter plot train', 'Dendrogram', 'Scatter plot test', 'PCA variance', 'Cross-Tabulation'])

    clusterization.base_model(X, ws, diagrams_dir=diagrams_folder_path)

    model_options = [
        'agglomerative'
    ]

    data_preprocessing = [
        {'scaling':False,
         'pca':False,
         'tsne': False,
         'truncatedsvd':True},
    ]

    n_components = np.append(np.arange(1, min(X.columns.size, 10), 1), X.columns.size)
    print(f'{n_components=}')

    for name in model_options:
        for preprocessing in data_preprocessing:
            for n_component in n_components:
                scaling = preprocessing['scaling']
                pca = preprocessing['pca']
                tsne = preprocessing['tsne']
                truncatedsvd = preprocessing['truncatedsvd']
                
                clusterize(X=X, name=name, real_labels=real_labels,n_features=n_component, scaling=scaling, pca=pca, tsne=tsne, truncatedsvd=truncatedsvd, ws=ws, diagrams_dir=diagrams_folder_path)

    wb.save(filename)

if __name__ == '__main__':
    main()