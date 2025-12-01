from sklearn.cluster import DBSCAN
import hdbscan
from sklearn.preprocessing import StandardScaler
import numpy as np
from sklearn.pipeline import Pipeline
from sklearn.cluster import AgglomerativeClustering
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from sklearn.metrics import silhouette_score
from scipy.cluster.hierarchy import dendrogram, linkage
from task01_data import points, new_points
from sklearn.model_selection import RandomizedSearchCV
from sklearn.model_selection import PredefinedSplit

global base_model_silhouette

def base_model(X, ws, diagrams_dir = './diagrams'):
    global base_model_silhouette

    labels = np.zeros(X.shape[0])

    inertia = None  # only for KMeans / MiniBatchKMeans
    base_model_silhouette = silhouette_score(X, labels) if len(set(labels)) > 1 else -1


    ws.append([
            f"Base model",
            '',
            X.shape[1],
            '' ,
            inertia if inertia is not None else "",
            base_model_silhouette if base_model_silhouette is not None else "",
            0
        ])
    
    
    
    plt.figure(figsize=(7, 5))
    plt.scatter(X.iloc[:, 0], X.iloc[:, 1], c=labels)
    plt.title(f"Base model — Cluster Scatterplot")
    plt.xlabel(X.columns[0])
    plt.ylabel(X.columns[1])
    plt.tight_layout()

    scatter_path = f"{diagrams_dir}/base_model_scatter.png"
    plt.savefig(scatter_path)
    plt.cla()

    img = openpyxl.drawing.image.Image(scatter_path)
    row = ws.max_row
    col = 'J'
    cell_ref = f"{col}{row}"

    img_width, img_height = img.width, img.height
    ws.column_dimensions[col].width = img_width / 7
    ws.row_dimensions[row].height = img_height

    img.anchor = cell_ref
    ws.add_image(img)

    col = chr(ord(col) + 1)


def silhouette_scorer(estimator, X):
    labels = estimator.fit_predict(X)
    if len(set(labels)) <= 1:   # silhouette cannot be computed
        return -1
    return silhouette_score(X, labels)

def clusterize(X, name, y=None, scaling=True, args = None, export_to_file = True, ws = None, diagrams_dir='./diagrams'):
    global base_model_silhouette

    param_grid = None
    reg = None
    
    if scaling == True:
        steps = [('scaler', StandardScaler())]
    else:
        steps = []


    if 'hdbscan' in name:
        if args == None:
            param_grid = {
                'hdbscan__min_cluster_size': np.arange(1, 50, 1)
            }
        else:
            param_grid = args.params
        reg = hdbscan.HDBSCAN()
        steps.append(('hdbscan', reg))
    elif 'dbscan' in name:
        if args == None:
            param_grid = {
                'dbscan__eps': np.arange(.1, 1, .05),
                'dbscan__min_samples': np.arange(1, 50, 1)
                }
        else:
            param_grid = args.params
        reg = DBSCAN()
        steps.append(('dbscan', reg))
    elif 'agglomerative' in name:
        if args == None:
            param_grid = {
                'agglomerative__linkage': ['single', 'average', 'complete'],
                'agglomerative__distance_threshold':np.arange(.1, 2, .05)
                }
        else:
            param_grid = args.params
        reg = AgglomerativeClustering(
            distance_threshold=1,
            n_clusters=None,
            compute_full_tree=True
        )

        steps.append(('agglomerative', reg))


    pipeline = Pipeline(steps)


    cv = [(np.arange(len(X)), np.arange(len(X)))]

    cv = RandomizedSearchCV(
        pipeline,
        param_distributions=param_grid,
        scoring=silhouette_scorer,
        cv=cv,
        n_iter=30
    )


    cv.fit(X=X)

    if export_to_file:
        pipeline = cv.best_estimator_

        clusterer = pipeline.named_steps[list(pipeline.named_steps.keys())[-1]]
        labels = pipeline.fit_predict(X)
        

        # Calculate metrics
        inertia = getattr(clusterer, "inertia_", None)  # only for KMeans / MiniBatchKMeans
        silhouette = silhouette_score(X, labels) if len(set(labels)) > 1 else None

        labels_y = pipeline.fit_predict(y)
        silhouette_test = silhouette_score(y, labels_y) if len(set(labels_y)) > 1 else None
        # ----------- SCATTERPLOT OF CLUSTERS -----------
        plt.figure(figsize=(7, 5))
        plt.scatter(X.iloc[:, 0], X.iloc[:, 1], c=labels)
        plt.title(f"{name} — Cluster Scatterplot")
        plt.xlabel(X.columns[0])
        plt.ylabel(X.columns[1])
        plt.tight_layout()

        scatter_path = f"{diagrams_dir}/{name}_scatter.png"
        plt.savefig(scatter_path)
        plt.cla()

        # ----------- SCATTERPLOT OF CLUSTERS -----------
        plt.figure(figsize=(7, 5))
        plt.scatter(y.iloc[:, 0], y.iloc[:, 1], c=labels_y)
        plt.title(f"{name} — Cluster Test Scatterplot")
        plt.xlabel(y.columns[0])
        plt.ylabel(y.columns[1])
        plt.tight_layout()

        scatter_path_test = f"{diagrams_dir}/{name}_scatter_test.png"
        plt.savefig(scatter_path_test)
        plt.cla()

        # ----------- DENDROGRAM (ONLY FOR HIERARCHICAL) -----------
        dendro_path = None
        if clusterer.__class__.__name__.startswith("Agglomerative"):
            plt.figure(figsize=(10, 5))
            Z = linkage(X, method=clusterer.linkage)
            dendrogram(Z)
            plt.title(f"{name} — Dendrogram ({clusterer.linkage} linkage)")
            plt.tight_layout()

            dendro_path = f"{diagrams_dir}/{name}_dendrogram.png"
            plt.savefig(dendro_path)
            plt.cla()

        # ----------- EXPORT TO EXCEL -----------
        col = 'J'
        for img_path in [scatter_path, dendro_path, scatter_path_test]:
            if not img_path:
                col = chr(ord(col) + 1)
                continue

            img = openpyxl.drawing.image.Image(img_path)
            row = ws.max_row + 1
            cell_ref = f"{col}{row}"

            img_width, img_height = img.width, img.height
            ws.column_dimensions[col].width = img_width / 7
            ws.row_dimensions[row].height = img_height

            img.anchor = cell_ref
            ws.add_image(img)

            col = chr(ord(col) + 1)

        # ----------- APPEND SUMMARY ROW -----------
        ws.append([
            f"{name} Clustering on {X.columns.tolist()}",
            scaling,
            X.shape[1],
            str(cv.best_params_) ,
            inertia if inertia is not None else "",
            silhouette if silhouette is not None else "",
            silhouette/base_model_silhouette * 100 - 100 if silhouette is not None else 0,
            silhouette_test if silhouette_test is not None else "",
            silhouette_test/base_model_silhouette * 100 - 100 if silhouette_test is not None else 0,


        ])

    return pipeline


def main():
    X = pd.DataFrame(points, columns=['X', 'Y'])
    y = pd.DataFrame(new_points, columns=['X', 'Y'])
    filename = './model_report_task01.xlsx'
    diagrams_folder_path = './diagrams'

    wb = openpyxl.Workbook()
    wb.create_sheet('ModelReport')
    ws = wb['ModelReport']

    ws.append(['Model', 'Scaling', 'Number of variables','Hyperparams', 'Inertia', 'Silhouette', 'Silhouette Increase from base model %', 'Silhouette for test data', 'Silhouette for test data increase %','Scatter plot train', 'Dendrogram', 'Scatter plot test'])

    base_model(X, ws)

    model_options = [
        'agglomerative',
        'dbscan',
        'hdbscan'
    ]

    for name in model_options:
        for scaling in [True, False]:
            clusterize(X=X, y=y, name=name, scaling=scaling, ws=ws)

    wb.save(filename)


if __name__ == '__main__':
    main()