import pandas as pd
from nltk.corpus import stopwords
import string
from gensim.corpora.dictionary import Dictionary
from nltk.tokenize import word_tokenize, regexp_tokenize
from gensim.models import TfidfModel
from collections import defaultdict
import statistics
from sklearn.model_selection import train_test_split, cross_val_score, KFold, GridSearchCV, RandomizedSearchCV
import numpy as np
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline
from sklearn.metrics import ConfusionMatrixDisplay, RocCurveDisplay, classification_report, roc_curve, recall_score, precision_score
import matplotlib.pyplot as plt
import openpyxl

def bow_to_words(bow, number, dictionary):
    words = bow[:number]
    words = list(map(lambda x : x[0], words))
    words = list(map(lambda x : dictionary[x],words))
    return words

def combine_bows(bows):
    combined = defaultdict(int)
    for bow in bows:
        for token_id, count in bow:
            combined[token_id] += count
    return list(combined.items())

base_model_precision = None
base_model_recall = None
base_model_f1 = None

def base_model(y, ws):
    global model_number
    global base_model_f1
    global base_model_recall
    global base_model_precision

    pred = 1

    y_pred = np.ones_like(y) * pred

    base_model_recall = recall_score(y, y_pred)
    base_model_precision = precision_score(y, y_pred)
    base_model_f1 = 2*base_model_precision*base_model_recall/(base_model_precision + base_model_recall + 1e-6)

    ws.append([f'Base model','', '', '', base_model_precision, 0, base_model_recall, 0, base_model_f1, 0])

def train_model(X, y, name, ws = None, args = None, export_to_file = True, diagrams = './diagrams',   base_model_f1 = base_model_f1,
                base_model_recall = base_model_recall,
                base_model_precision = base_model_precision):
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=21, stratify=y)

    n_splits = 5
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)
    
    param_grid = None
    steps = []

    if 'count' in name:

        # because i will give it the tokens myself
        reg = CountVectorizer(
            tokenizer=lambda x: x,
            preprocessor=lambda x: x,
            token_pattern=None)
        
        steps.append(('count', reg))

    elif 'tfidf' in name:
        reg = TfidfVectorizer(
            tokenizer=lambda x: x,
            preprocessor=lambda x: x,
            token_pattern=None
        )
        steps.append(('tfidf', reg))

    param_grid = {
        'nb__alpha': [0.1, 0.5, 1.0, 2.0],
        'nb__fit_prior': [True, False]
    }


    naivebayes = MultinomialNB()
    steps.append(('nb', naivebayes))

    pipeline = Pipeline(steps)
    
    cv = RandomizedSearchCV(pipeline, param_distributions=param_grid,cv=kf,return_train_score=True, scoring='f1')
    cv.fit(X=X_train, y=y_train)

    y_pred = cv.predict(X_test)

    if export_to_file:
        
        ConfusionMatrixDisplay.from_predictions(y_test, y_pred)
        plt.tight_layout()
        if hasattr(X, 'loc'):
            confusion_matrix_path = f'{diagrams}/{name}_confusion_matrix_{X_test.columns[0]}.png'
        else:
            confusion_matrix_path = f'{diagrams}/{name}_confusion_matrix.png'

        plt.savefig(confusion_matrix_path)
        plt.cla()

        row_num = ws.max_row + 1

        img = openpyxl.drawing.image.Image(confusion_matrix_path)
        cell_ref = f'K{row_num}'

        img_width, img_height = img.width, img.height

        column_letter = 'K'
        col_width = img_width / 7
        row_height = img_height / 1.333

        ws.column_dimensions[column_letter].width = col_width
        ws.row_dimensions[row_num].height = row_height

        img.anchor = cell_ref
        ws.add_image(img, cell_ref)

        scores = classification_report(y_test, y_pred, labels=[1], output_dict=True)

        if isinstance(cv, Pipeline):
            ws.append([f'{name} model','', '', '', scores['1']['precision'], scores['1']['precision']/base_model_precision * 100 - 100,scores['1']['recall'], scores['1']['recall']/base_model_recall * 100 - 100, scores['1']['f1-score'], scores['1']['f1-score']/base_model_f1 * 100 -100])
        else:
            ws.append([f'{name} model','','', str(cv.best_params_), scores['1']['precision'], scores['1']['precision']/base_model_precision * 100 - 100,scores['1']['recall'], scores['1']['recall']/base_model_recall * 100 - 100, scores['1']['f1-score'], scores['1']['f1-score']/base_model_f1 * 100 -100])

    return cv, [y_test, y_pred], scores

def extract_top_words_from_nb(best_estimator, top_n=10):
    # get vectorizer
    for name, step in best_estimator.named_steps.items():
        if isinstance(step, (CountVectorizer, TfidfVectorizer)):
            vectorizer = step
        if isinstance(step, MultinomialNB):
            nb = step

    feature_names = np.array(vectorizer.get_feature_names_out())

    fake_scores = nb.feature_log_prob_[1] - nb.feature_log_prob_[0]
    real_scores = nb.feature_log_prob_[0] - nb.feature_log_prob_[1]

    top_fake_idx = np.argsort(fake_scores)[-top_n:][::-1]
    top_real_idx = np.argsort(real_scores)[-top_n:][::-1]

    top_fake_words = feature_names[top_fake_idx]
    top_real_words = feature_names[top_real_idx]

    return top_fake_words, top_real_words



def main():
    df = pd.read_csv('fake_or_real_news.csv')
    df['label'] = df['label'].map(lambda x : 1 if x=='FAKE' else 0)

    filename = "model_report.xlsx"

    wb = openpyxl.Workbook()
    wb.create_sheet('ModelReport')
    ws = wb['ModelReport']

    corpus = df['text']

    stop_words = set(stopwords.words('english'))
    punct = set(string.punctuation)
    punct.add("''")
    punct.add("``")
    punct.add("||")


    tokenized = [word_tokenize(doc.lower()) for doc in corpus]
    
    ws.append(['Model', 'Scaling', 'Number of variables','Hyperparams', "Precision", "Precision increase from base model (in %)", "Recall", "Recall increase from base model (in %)", "F1 Score", "F1 score increase from base model (in %)", "Confusion matrix", "Comments"])

    base_model(df['label'], ws)

    print("base model done!")

    best_model = None
    best_f1 = -1


    model_names = ['count', 'tfidf']
    for name in model_names:
        cv, (y_test, y_pred), scores = train_model(tokenized, df['label'], name, ws, base_model_recall=base_model_recall,
                    base_model_precision=base_model_precision,
                    base_model_f1=base_model_f1)
        
        f1 = scores['1']['f1-score']

        if f1 > best_f1:
            best_f1 = f1
            best_model = cv.best_estimator_
        


    print("models on raw data trained!")

    tokenized_without_punct_stop_words = [[t for t in doc 
                   if t not in stop_words and t not in punct] 
                    for doc in tokenized]
    
    for name in model_names:
        cv, (y_test, y_pred), scores = train_model(tokenized_without_punct_stop_words, df['label'], name, ws, base_model_recall=base_model_recall,
                    base_model_precision=base_model_precision,
                    base_model_f1=base_model_f1)
        
        f1 = scores['1']['f1-score']

        if f1 > best_f1:
            best_f1 = f1
            best_model = cv.best_estimator_

    print("models on clean data trained!")

    

    top_fake, top_real = extract_top_words_from_nb(best_model, top_n=10)

    ws_words = wb.create_sheet('BestModelWords')

    ws_words.append(['FAKE News Words', 'REAL News Words'])

    for f, r in zip(top_fake, top_real):
        ws_words.append([f, r])

    wb.save(filename)



if __name__ == '__main__':
    main()