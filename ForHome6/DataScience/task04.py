from sklearn.linear_model import LinearRegression, Ridge, Lasso
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split, cross_val_score, KFold, GridSearchCV, RandomizedSearchCV
import openpyxl
import numpy as np
from sklearn.metrics import r2_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import VotingRegressor

model_number = 1
base_model_score = None
feature_importance_diagram_created = False

def base_model( y, ws):
    global model_number
    global base_model_score

    y_train, y_test = train_test_split( y, test_size=0.3, random_state=21)
    pred = np.mean(y_train)

    y_pred = np.ones_like(y_test) * pred
    base_model_score = r2_score(y_test, y_pred) + 1e-6

    ws.append([f'Base model', '', '', base_model_score, 0])
    model_number += 1
    

def train_and_test_model(X, y, ws, name, args=None, export_to_file = True):

    global model_number
    global base_model_score
    global feature_importance_diagram_created

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=21)


    n_splits = 5
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)
    
    param_grid = None
    reg = None
    
    steps = [('scaler', StandardScaler())]

    if "linear" in name:
        param_grid = {}
        reg = LinearRegression()
        steps.append(('linear', reg))
    elif "ridge" in name:
        param_grid = {
            'ridge__alpha': np.arange(.1, 1, .1),
            'ridge__solver': ['sag', 'lsqr']
        }
        reg = Ridge()
        steps.append(('ridge', reg))
    elif "lasso" in name:
        param_grid = {
            'lasso__alpha': np.arange(.1, 1, .1)
        }
        reg = Lasso()
        steps.append(('lasso', reg))
    elif "cart" in name:
        param_grid = {
            'cart__criterion':['squared_error','absolute_error'],
            'cart__max_depth': np.arange(1, 100, 1),
            'cart__min_samples_split': np.arange(2, 50, 1)
        }
        reg = DecisionTreeRegressor()
        steps.append(('cart', reg))
    elif 'ensemble' in name:
        if args == None or (not ('models' in args)):
            models = ['svm']
        else:
            models = args['models']


        if args == None or (not ('params' in args)) :
            param_grid = {}
        else:
            param_grid = args['params']

        trained_models = []
        for model in models:
            cv = train_and_test_model(X, y, ws, model, export_to_file=False)
            trained_models.append((model, cv))

        reg = VotingRegressor(trained_models)
        steps.append(('ensemble', reg))
    else:
        raise Exception("Invalid model name!")
    
    pipeline = Pipeline(steps)

    cv = RandomizedSearchCV(pipeline, param_distributions=param_grid,cv=kf, return_train_score=True, scoring='r2')
    cv.fit(X=X_train, y=y_train)

    if export_to_file:
        results = pd.DataFrame(cv.cv_results_)

        results['params_label'] = results['params'].apply(
            lambda p: ', '.join(f'{k}={v}' for k, v in p.items())
        )

        results = results.sort_values(by='mean_test_score', ascending=True)

        y_labels = results['params_label']
        train_scores = results['mean_train_score']
        test_scores = results['mean_test_score']

        plt.figure(figsize=(10, 5))
        bar_width = 0.4
        y_positions = range(len(y_labels))

        plt.barh(
            [y - bar_width/2 for y in y_positions],
            train_scores,
            height=bar_width,
            label='Train score',
        )
        plt.barh(
            [y + bar_width/2 for y in y_positions],
            test_scores,
            height=bar_width,
            label='Test score',
        )

        plt.yticks(y_positions, y_labels)
        plt.xlabel("Score")
        plt.title("Train vs Test Score for Each Hyperparameter Combination")
        plt.legend()

        plt.tight_layout()
        plt.savefig(f'./diagrams/{name}_regression_diagram_{X_test.columns[0]}.png')
        plt.cla()

        img = openpyxl.drawing.image.Image(f'diagrams/{name}_regression_diagram_{X_test.columns[0]}.png')
        

        cell_ref = f'F{1 + model_number}'

        img_width, img_height = img.width, img.height

        column_letter = 'F'
        col_width = img_width / 7
        row_height = img_height / 1.333


        ws.column_dimensions[column_letter].width = col_width
        ws.row_dimensions[1 + model_number].height = row_height

        img.anchor = cell_ref
        ws.add_image(img, cell_ref)

        if 'lasso' in name and not feature_importance_diagram_created:
            best = cv.best_estimator_
            lasso = best.named_steps["lasso"]
            coefs = lasso.coef_
            feature_importance = pd.DataFrame({
                "feature": X.columns,
                "coef": coefs,
                "importance": np.abs(coefs)
            })
            feature_importance = feature_importance.sort_values("importance", ascending=False)

            plt.figure(figsize=(10, 6))
            plt.barh(feature_importance["feature"], feature_importance["importance"])
            plt.gca().invert_yaxis()
            plt.xlabel("Importance (|coefficient|)")
            plt.title("Lasso Feature Importance")
            plt.tight_layout()
            plt.savefig(f'./diagrams/feature_importance.png')
            plt.cla()
            feature_importance_diagram_created = True

        ws.append([f'{name} Regression on {X.columns}', X.shape[1], str(cv.best_params_), cv.best_score_, cv.best_score_/base_model_score * 100 - 100])
        model_number += 1


    return cv


def main():
    df = pd.read_csv('auto.csv')
    df_dummies = pd.get_dummies(df['origin'], drop_first=True, dtype=int)
    df = pd.concat([df_dummies,df.drop(columns=['origin'])], axis=1)

    filename = './model_report_task04.xlsx'

    wb = openpyxl.Workbook()
    wb.create_sheet('ModelReport')
    ws = wb['ModelReport']

    ws.append(['Model', 'Number of variables','Hyperparams', 'R2 score', "R2 score % increase from base model"])
    y = df['mpg']
    df = df.drop(columns=['mpg'])

    
    base_model(y, ws)


    X_options = [
        df,
        df[['US','Europe','accel']],
        df[['US','Europe', 'accel', 'size']]
    ]

    model_options = [
        'linear',
        'ridge', 
        'lasso',
        'cart',
        'ensemble'
    ] 

    ensemble_models = [
        ['linear', 'ridge', 'lasso', 'cart'],
        ['cart', 'lasso'],
        ['cart1', 'cart2']
    ]
    
    for X in X_options:
        for model in model_options:
            if model == 'ensemble':
                for i in range(len(ensemble_models)):
                    args = {'models': ensemble_models[i]}
                    train_and_test_model(X, y, wb['ModelReport'], model,args=args)
            else:
                train_and_test_model(X, y, wb['ModelReport'], model)


    img = openpyxl.drawing.image.Image(f'diagrams/feature_importance.png')
        
    row = ws.max_row + 1
    cell_ref = f'A{row}'

    img_width, img_height = img.width, img.height

    column_letter = 'A'
    img.anchor = cell_ref
    ws.add_image(img, cell_ref)

    wb.save(filename)
    

if __name__ == '__main__':
    main()